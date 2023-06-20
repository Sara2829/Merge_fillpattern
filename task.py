import openpyxl
import pandas as pd
import openpyxl as op
from openpyxl import workbook
from openpyxl.styles import PatternFill
#setting path for source and destination files
source_file = pd.read_excel('source.xlsx')
destination_file= pd.read_excel('destination.xlsx')
#reading one coloumn
source_coloumn =pd.read_excel('source.xlsx',usecols=['subject'])
#loading workbook 
wb = openpyxl.load_workbook('destination.xlsx')
sheet = wb.active
#merging cells
sheet.merge_cells('A1:A4')
sheet.cell(row=1,column=1).value= 'merge cells '

#filling color
sheet['A1']= PatternFill(patternType='solid', fgColor= '0A6EBD' )
wb.save('destination.xlsx')
